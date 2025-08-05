import os
import logging
from io import BytesIO
from math import ceil

import openpyxl
import pandas as pd
import xlwings as xw
import copy
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter
from .transform_excel import xls2xlsx
from openpyxl.worksheet.pagebreak import Break
from .enhanced_height_calculator import get_height_calculator

# --- 框架核心：通用工具函数 ---

# 精度安全的twip计算工具
MM_PER_PT = 0.3527777778          # 精确转换常数，不要截断
TWIP = 20                         # 1 pt = 20 twip
GRID_TWIP = 15                   # thin border ≈ 0.75 pt，相邻行共用

def mm_to_twip(mm: float) -> int:
    """直接从毫米转换为twip整数，避免pt中间截断"""
    return round(mm / MM_PER_PT * TWIP)

def pt_to_twip(pt: float) -> int:
    """点转换为twip整数"""
    return round(pt * TWIP)

def twip_to_pt(twip: int) -> float:
    """twip转换回点（仅用于日志显示）"""
    return twip / TWIP

def mm_to_points(mm):
    """将毫米转换为点（保留向后兼容）。"""
    return round(mm * (1 / 0.35278), 2)


def points_to_mm(points):
    """将点转换为毫米。"""
    return round(points * 0.35278, 2)


def inch_to_mm(inch):
    """将英寸转换为毫米。"""
    return round(inch * 25.4, 2)


def cleanup_stream(stream):
    """清除内存流。"""
    if stream:
        stream.close()


def get_cell_or_merged_width(sheet, cell_address):
    """
    获取指定单元格的列宽，支持合并单元格。
    """
    for merged_range in sheet.merged_cells.ranges:
        if cell_address in merged_range:
            min_col, _, max_col, _ = merged_range.bounds
            total_width = 0
            for col_idx in range(min_col, max_col + 1):
                column_letter = get_column_letter(col_idx)
                dim = sheet.column_dimensions.get(column_letter)
                if dim and dim.width:
                    total_width += dim.width
            return total_width

    col_letter = "".join(filter(str.isalpha, cell_address))
    dim = sheet.column_dimensions.get(col_letter)
    if dim:
        return dim.width
    return 0


def get_subset(data, start_key, end_key):
    """
    根据给定的起始和结束键值，从列表中获取子集。
    """
    if not isinstance(data, list):
        data = data.tolist()

    start_index = 0
    if start_key != "" and start_key is not None:
        try:
            start_index = data.index(start_key)
        except ValueError:
            logging.warning(f"未找到指定的起点: {start_key}, 将从头开始。")
            start_index = 0

    end_index = len(data)
    if end_key != "" and end_key is not None:
        try:
            end_index = data.index(end_key) + 1
        except ValueError:
            logging.warning(f"未找到指定的终点: {end_key}, 将持续到末尾。")
            end_index = len(data)

    start_index = max(0, start_index)
    end_index = min(len(data), end_index)

    return data[start_index:end_index]


def get_one_row_height(rng):
    """
    使用xlwings计算自适应的单行文本高度。
    """
    original_value = rng.value
    rng.value = "宋"
    rng.autofit()
    # 设置一个只能容纳单个字符的宽度
    rng.column_width = rng.column_width
    rng.value = 10 * "宋"
    rng.api.WrapText = True
    one_row_height = rng.row_height / 10  # 取平均行高
    rng.value = original_value
    rng.api.WrapText = False
    rng.autofit()
    return one_row_height


def set_autofit_fontsize(rng, current_col, current_row, sheet):
    """
    通过调整字体大小和行高，使单元格内容完全显示。
    支持三种行高计算方案：xlwings（原始）、gdi、pillow
    """
    rangeStr = f"{get_column_letter(current_col)}{current_row}"

    # 如果未明确设置，则使用默认行高
    row_height = sheet.row_dimensions[current_row].height
    if row_height is None:
        row_height = 64  # 默认行高（磅），适用于11号字体。
    
    column_width = get_cell_or_merged_width(sheet, rangeStr)

    if column_width is None or column_width == 0:
        logging.warning(f"单元格 {rangeStr} 列宽为0或未定义，跳过自适应调整。")
        return

    # 获取单元格文本内容
    cell_text = sheet[rangeStr].value
    if not cell_text:
        return
    
    # 获取增强行高计算器
    height_calculator = get_height_calculator()
    current_method = height_calculator.method
    
    # 移除调试信息以提升性能
    
    try:
        if current_method == 'xlwings':
            # 使用原始xlwings方案
            rng.value = cell_text
            rng.autofit()
            
            adjust_row_height = ceil(
                (rng.column_width * 1.15) / column_width
            ) * get_one_row_height(rng)
            
            # 迭代减小字体大小直到内容适应，或达到最小11号字
            current_font_size = rng.font.size
            while row_height < adjust_row_height and current_font_size > 11:
                current_font_size -= 1
                rng.font.size = current_font_size
                rng.autofit()
                adjust_row_height = ceil(
                    (rng.column_width * 1.06) / column_width
                ) * get_one_row_height(rng)
            
            # 如果字体已是最小，但内容仍不适应，则调整行高
            if row_height < adjust_row_height:
                sheet.row_dimensions[current_row].height = adjust_row_height
            else:
                sheet.row_dimensions[current_row].height = (
                    row_height if row_height is not None else 64
                )
            
            # 将调整后的字体大小应用于最终的sheet
            sheet[rangeStr].font = Font(size=current_font_size)
            
        else:
            # 使用GDI或Pillow方案进行精确计算
            calculated_height = height_calculator.calculate_height(
                rng, str(cell_text), column_width, f"第{current_row}行{rangeStr}: "
            )
            
            # 重要：确保计算出的行高不低于原始行高（保持原始逻辑）
            final_height = max(calculated_height, row_height)
            
            # 应用最终行高
            sheet.row_dimensions[current_row].height = final_height
            
            # 保持字体大小为11号（这些方案专门为11号字体优化）
            sheet[rangeStr].font = Font(size=11)
            
    except Exception as e:
        logging.error(f"单元格 {rangeStr} 行高计算失败 ({current_method}方案): {e}")
        # 回退到默认行高
        sheet.row_dimensions[current_row].height = row_height if row_height is not None else 64
        sheet[rangeStr].font = Font(size=11)

    if sheet[rangeStr].font.size < 11:
        logging.info(
            f"内容自适应调整: 单元格 {rangeStr}, "
            f"最终字号 {sheet[rangeStr].font.size}, "
            f"行高 {sheet.row_dimensions[current_row].height}"
        )

    # 为下一次计算重置rng的字体大小
    rng.font.size = 11


# --- 框架核心：主要逻辑 ---


def load_data(excel_path):
    """
    加载Excel数据，自动处理.xls和.xlsx格式。
    """
    if not os.path.exists(excel_path):
        logging.error(f"文件路径不存在: {excel_path}")
        return None

    filename = os.path.basename(excel_path)
    if filename.lower().endswith(".xls"):
        logging.info(f"检测到.xls文件: {filename}，正在尝试转换为.xlsx...")
        try:
            xls2xlsx(excel_path)
            excel_path = os.path.splitext(excel_path)[0] + ".xlsx"
            logging.info(f"转换成功，将从 {os.path.basename(excel_path)} 加载数据。")
        except Exception as e:
            logging.error(f"转换 {filename} 失败: {e}")
            return None

    try:
        data = pd.read_excel(excel_path)
        return data
    except Exception as e:
        logging.error(f"加载数据时出错: {e}")
        return None


def prepare_template(template_path):
    """
    以内存流方式加载模板文件。
    """
    try:
        with open(template_path, "rb") as file:
            stream = BytesIO(file.read())
        stream.seek(0)
        return stream
    except FileNotFoundError:
        logging.error(f"模板文件不存在: {template_path}")
        return None
    except Exception as e:
        logging.error(f"加载模板时出错: {e}")
        return None


def generate_one_archive_directory(
    archive_data,
    template_stream,
    output_folder,
    archive_id,
    rng_for_calc,
    index,
    column_mapping,
    autofit_columns,
    static_cells={},
    title_row_num=4,
    page_height_mm=296,
    direct_print=False,
    printer_name=None,
    print_copies=1,
    cancel_flag=None,
):
    """
    为单个案卷生成目录，处理分页和内容自适应。
    """
    if not template_stream:
        logging.error("模板流无效，无法生成目录。")
        return 0

    template_stream.seek(0)  # 每次使用时重置流指针
    new_wb = openpyxl.load_workbook(template_stream)
    sheet = new_wb.worksheets[0]

    if str(sheet.page_setup.paperSize) == "9":  # A4 paper code
        # 修复纸张方向判断逻辑
        if sheet.page_setup.orientation == 'landscape':
            page_height_mm = 209  # A4横向：宽度变成高度
        else:
            page_height_mm = 296  # A4纵向：正常高度
    elif sheet.page_setup.paperSize:
        logging.warning(
            f"模板纸张大小 {sheet.page_setup.paperSize}, 非A4, 高度计算可能不准."
        )

    # --- 页面尺寸和边距计算 ---

    margins = sheet.page_margins
    top_margin_mm = inch_to_mm(margins.top)
    bottom_margin_mm = inch_to_mm(margins.bottom)
    
    # 计算实际可用内容高度（使用全twip精度计算）
    # 注意：需要扣除页脚边距，与WPS预览完全一致
    paper_h_twip = mm_to_twip(page_height_mm)
    top_twip = mm_to_twip(inch_to_mm(margins.top))
    bottom_twip = mm_to_twip(inch_to_mm(margins.bottom))
    footer_twip = mm_to_twip(inch_to_mm(margins.footer))  # Excel默认页脚距底端1.27cm
    
    # 添加缩放系数支持
    scale = (sheet.page_setup.scale or 100) / 100  # 0-400%
    content_height_twip = int((paper_h_twip - top_twip - bottom_twip - footer_twip) * scale)
    
    # 转换为点用于向后兼容（但分页判断使用twip）
    content_height_points = twip_to_pt(content_height_twip)
    
    # 生产环境移除详细的页面高度计算日志

    # --- 填充静态信息 ---
    for cell, value in static_cells.items():
        if value and not pd.isna(value):
            sheet[cell] = value

    # --- 模板行 ---
    template_data_start_row_num = title_row_num + 1
    if template_data_start_row_num <= title_row_num:
        # logging.warning(
        #     f"模板数据起始行号 ({template_data_start_row_num}) 无效. 使用默认值 {title_row_num + 1}."
        # )
        template_data_actual_ref_row = title_row_num + 1
    else:
        template_data_actual_ref_row = template_data_start_row_num

    default_data_row_height = sheet.row_dimensions[template_data_actual_ref_row].height
    if default_data_row_height is None:
        default_data_row_height = 64
    #     logging.warning(
    #         f"模板行 {template_data_actual_ref_row} 未设置明确行高，用默认值 {default_data_row_height} points."
    #     )
    # logging.info(
    #     f"模板样式从第 {template_data_actual_ref_row} 行获取, 默认数据行高: {default_data_row_height} points"
    # )
    print_title_row_height = sum(
        (
            sheet.row_dimensions[r].height
            if sheet.row_dimensions[r].height is not None
            else 64
        )
        for r in range(1, title_row_num + 1)
    )
    sheet.print_title_rows = f"1:{title_row_num}"

    template_row_styles = []
    for col_idx in range(1, sheet.max_column + 1):
        tc = sheet.cell(row=template_data_actual_ref_row, column=col_idx)
        template_row_styles.append(
            {
                "font": copy.copy(tc.font),
                "border": copy.copy(tc.border),
                "fill": copy.copy(tc.fill),
                "alignment": copy.copy(tc.alignment),
                "number_format": tc.number_format,
                "protection": copy.copy(tc.protection),
            }
        )

    # --- 填充动态表格数据 ---
    current_row = title_row_num + 1
    current_height_points = print_title_row_height * scale  # 标题行高度也需要应用缩放
    curr_height_twip = pt_to_twip(print_title_row_height * scale)  # twip精度变量
    pages = []
    page_count = 0
    page_start_row = current_row

    for row_idx, (_, row_data) in enumerate(archive_data.iterrows()):
        # 填充一行数据
        for col_idx_m1, style_to_apply in enumerate(template_row_styles):
            cell = sheet.cell(row=current_row, column=col_idx_m1 + 1)
            cell.font = copy.copy(style_to_apply["font"])
            cell.border = copy.copy(style_to_apply["border"])
            cell.fill = copy.copy(style_to_apply["fill"])
            cell.alignment = copy.copy(style_to_apply["alignment"])
            cell.number_format = style_to_apply["number_format"]
            cell.protection = copy.copy(style_to_apply["protection"])

        for dest_col, src_col in column_mapping.items():
            value = row_data.get(src_col, "")
            # 确保值不是NaN，并对字符串进行strip操作
            if pd.isna(value):
                value = ""
            if isinstance(value, str):
                value = value.strip()
            sheet.cell(row=current_row, column=dest_col, value=value)

        # 对指定列进行内容自适应调整
        for col_index in autofit_columns:
            # 在自适应前检查是否为NaN
            src_col_for_check = column_mapping.get(col_index)
            if src_col_for_check and pd.notna(row_data.get(src_col_for_check)):
                set_autofit_fontsize(rng_for_calc, col_index, current_row, sheet)

        # 获取行高并转换为twip进行精确计算
        row_height = (
            sheet.row_dimensions[current_row].height
            if sheet.row_dimensions[current_row].height is not None
            else 64
        )
        
        # 使用全twip精度分页逻辑
        row_twip = pt_to_twip(row_height * scale)
        left_twip = content_height_twip - curr_height_twip
        
        # 向后兼容的点值计算（仅用于日志）
        scaled_row_height = row_height * scale
        left_points = content_height_points - current_height_points
        
        # ① 先用完整行高判断是否需要分页（避免"判页用扣减值"的问题）
        need_break = (curr_height_twip + row_twip - 1 > content_height_twip 
                     and current_height_points > 0)
        
        if need_break:
            # 补丁①：若后面已无数据，直接跳出循环，避免插空白分页
            has_more_rows = (row_idx + 1) < len(archive_data)  # 检查是否还有后续行
            if not has_more_rows:
                # 最后一条数据：不要插入分页符，直接翻页状态即可
                pages.append((page_start_row, current_row - 1))
                page_start_row = current_row
                current_height_points = print_title_row_height * scale
                curr_height_twip = pt_to_twip(print_title_row_height * scale)
                page_count += 1
            else:
                # 正常情况：插入手动分页符
                sheet.row_breaks.append(Break(id=current_row - 1))
                logging.info(
                    f"页码分割于{current_row -1}行后,(案卷 {archive_id})-twip比较: {curr_height_twip}+{row_twip}-1 > {content_height_twip}, pt值: {current_height_points:.2f}/{content_height_points:.2f}, 当前行高 {scaled_row_height:.2f}(缩放{scale:.0%})"
                )
                pages.append((page_start_row, current_row - 1))
                page_start_row = current_row
                current_height_points = print_title_row_height * scale
                curr_height_twip = pt_to_twip(print_title_row_height * scale)
                page_count += 1
            
            # ② 分页后，当前行成为新页首行，不扣 GRID_TWIP
            effective_row_twip = row_twip
        else:
            # 仍在同页，根据是否为首行决定是否扣减
            is_first_row_on_page = (current_row == page_start_row)
            effective_row_twip = row_twip if is_first_row_on_page else row_twip - GRID_TWIP
            # 当前行不会导致超出，正常累加行高
            current_height_points += scaled_row_height
        
        # 无论是否分页，都要累加当前行的实际占用高度
        curr_height_twip += effective_row_twip

        current_row += 1
    if (
        page_count > 0
        and 0 < current_height_points < content_height_points
        and default_data_row_height > 0
    ):
        # 修复：使用twip精度计算填充空行数量，避免高估
        # ② 末页空行的行高计算 —— 同理首行不扣 GRID_TWIP
        is_first_fill_row = (current_row == page_start_row)          # 仍在新页开始处
        fill_row_twip = (
            pt_to_twip(default_data_row_height * scale)
            - (0 if is_first_fill_row else GRID_TWIP)
        )
        remaining_twip = content_height_twip - curr_height_twip
        num_fill = max(0, (remaining_twip - 1) // fill_row_twip)  # 留1twip安全余量
        
        # 向后兼容的点值计算（仅用于日志）
        remaining_h = content_height_points - current_height_points
        logging.info(
            f"案卷 {archive_id},末页(页 {page_count})数据高{current_height_points:.2f}/{content_height_points:.2f}.剩余{remaining_h:.2f}pt({remaining_twip}twip).填充 {num_fill}空行(每行{default_data_row_height:.2f}pt,实际{twip_to_pt(fill_row_twip):.2f}pt)."
        )
        for _ in range(num_fill):
            for col_idx_m1, style_to_apply in enumerate(template_row_styles):
                fill_cell = sheet.cell(row=current_row, column=col_idx_m1 + 1)
                fill_cell.font = copy.copy(style_to_apply["font"])
                fill_cell.border = copy.copy(style_to_apply["border"])
                fill_cell.fill = copy.copy(style_to_apply["fill"])
                fill_cell.alignment = copy.copy(style_to_apply["alignment"])
                fill_cell.number_format = style_to_apply["number_format"]
                fill_cell.protection = copy.copy(style_to_apply["protection"])
            sheet.row_dimensions[current_row].height = default_data_row_height
            current_row += 1
    pages.append((page_start_row, current_row - 1))
    
    # 补丁②：清理孤立分页符（使用正确的openpyxl API）
    def prune_trailing_breaks(ws, last_row_in_print_area: int):
        """删除所有位于最后有效行之后的手动分页符，避免空白尾页"""
        pb = ws.row_breaks
        if hasattr(pb, "brk") and pb.brk is not None:
            brks = list(pb.brk)  # tuple -> list
            pruned = [b for b in brks if b.id < last_row_in_print_area]
            if len(pruned) != len(brks):
                logging.info(f"清理分页符: 原有{len(brks)}个, 清理后{len(pruned)}个")
                pb.brk = tuple(pruned)  # 回写

    prune_trailing_breaks(sheet, current_row - 1)
    
    if column_mapping:
        last_col_letter = get_column_letter(max(column_mapping.keys()))
        sheet.print_area = f"A1:{last_col_letter}{current_row - 1}"

    # --- 保存文件 ---
    os.makedirs(output_folder, exist_ok=True)
    # 替换文件名中的非法字符
    safe_archive_id = "".join(
        c for c in archive_id if c.isalnum() or c in ("-", "_")
    ).rstrip()
    save_path = os.path.join(output_folder, f"{safe_archive_id}.xlsx")

    try:
        new_wb.save(save_path)
        logging.info(
            f"[{index:04d}] 目录已保存: {safe_archive_id}.xlsx, 共计 {len(pages)} 页"
        )
        
        # 边转换边打印模式
        if direct_print and printer_name:
            try:
                # 导入打印服务（避免循环导入）
                import sys
                import os as os_module
                # 添加项目根路径到sys.path
                current_dir = os_module.path.dirname(os_module.path.dirname(os_module.path.abspath(__file__)))
                if current_dir not in sys.path:
                    sys.path.insert(0, current_dir)
                
                from utils.print_service import get_print_service
                print_service = get_print_service()
                
                # 使用异步打印，不阻塞文件生成
                future = print_service.async_print(save_path, printer_name, print_copies)
                logging.info(f"[{index:04d}] 已提交异步打印任务: {safe_archive_id}.xlsx -> {printer_name}")
                
            except Exception as print_err:
                logging.error(f"[{index:04d}] 提交异步打印任务异常: {print_err}")
        
        return len(pages)
    except PermissionError:
        logging.warning(
            f"[{index:04d}] 权限错误: 无法保存 {save_path}。请确保文件未被打开。"
        )
    except Exception as e:
        logging.error(f"[{index:04d}] 保存文件时出错 {save_path}: {e}")
    finally:
        new_wb.close()

    return 0
