"""
测试配置和公共工具模块
为所有测试提供统一的配置、工具函数和基础类
"""

import os
import sys
import tempfile
import shutil
import pytest
import logging
from pathlib import Path
from typing import Dict, Any, Optional
import pandas as pd
from unittest.mock import MagicMock

# 添加项目根目录到Python路径
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# 测试配置
TEST_CONFIG = {
    'timeout': 30,  # 测试超时时间（秒）
    'temp_dir_prefix': 'adg_test_',
    'log_level': logging.WARNING,  # 测试时减少日志噪音
    'mock_excel': True,  # 是否模拟Excel操作
    'benchmark_iterations': 3,  # 性能测试迭代次数
}

class TestEnvironment:
    """测试环境管理器"""
    
    def __init__(self):
        self.temp_dir = None
        self.temp_files = []
        self.original_log_level = None
        
    def setup(self):
        """设置测试环境"""
        # 创建临时目录
        self.temp_dir = tempfile.mkdtemp(prefix=TEST_CONFIG['temp_dir_prefix'])
        
        # 设置日志级别
        self.original_log_level = logging.getLogger().level
        logging.getLogger().setLevel(TEST_CONFIG['log_level'])
        
        return self
    
    def cleanup(self):
        """清理测试环境"""
        # 清理临时文件
        for file_path in self.temp_files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass
        
        # 清理临时目录
        if self.temp_dir and os.path.exists(self.temp_dir):
            try:
                shutil.rmtree(self.temp_dir)
            except Exception:
                pass
        
        # 恢复日志级别
        if self.original_log_level is not None:
            logging.getLogger().setLevel(self.original_log_level)
    
    def create_temp_file(self, content: str = '', suffix: str = '.txt') -> str:
        """创建临时文件"""
        fd, path = tempfile.mkstemp(suffix=suffix, dir=self.temp_dir)
        try:
            with os.fdopen(fd, 'w', encoding='utf-8') as f:
                f.write(content)
        except Exception:
            os.close(fd)
            raise
        
        self.temp_files.append(path)
        return path
    
    def create_test_excel(self, data: Dict[str, list], filename: str = 'test.xlsx') -> str:
        """创建测试用的Excel文件"""
        filepath = os.path.join(self.temp_dir, filename)
        df = pd.DataFrame(data)
        df.to_excel(filepath, index=False)
        self.temp_files.append(filepath)
        return filepath

def create_mock_archive_data(num_records: int = 10) -> pd.DataFrame:
    """创建模拟档案数据"""
    return pd.DataFrame({
        '案卷档号': [f'ZYZS2023-Y-{i:04d}' for i in range(1, num_records + 1)],
        '文件名': [f'文件{i:02d}' for i in range(1, num_records + 1)],
        '页数': [f'{i}' for i in range(1, num_records + 1)],
        '备注': [f'备注{i}' if i % 3 == 0 else '' for i in range(1, num_records + 1)],
    })

def create_mock_template() -> bytes:
    """创建模拟Excel模板"""
    import io
    import openpyxl
    from openpyxl.styles import Font, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "目录模板"
    
    # 设置标题行
    headers = ['序号', '文件名', '页数', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(name='SimSun', size=11, bold=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 设置数据行模板
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = Font(name='SimSun', size=11)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    
    # 保存到字节流
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()

class MockHeightCalculator:
    """模拟行高计算器"""
    
    def __init__(self, method='xlwings'):
        self.method = method
        self.performance_stats = {
            'xlwings': {'count': 0, 'total_time': 0},
            'gdi': {'count': 0, 'total_time': 0},
            'pillow': {'count': 0, 'total_time': 0}
        }
    
    def calculate_height(self, rng, text: str, column_width: float, row_info: str = "") -> float:
        """模拟行高计算"""
        import time
        import random
        
        # 模拟计算时间
        time.sleep(random.uniform(0.001, 0.005))
        
        # 简单的行高计算逻辑
        if not text or not text.strip():
            return 16.0
        
        # 基于文本长度和列宽估算行数
        char_width = 7  # 平均字符宽度（像素）
        available_width = column_width * char_width
        if available_width <= 0:
            available_width = 1  # 避免除零错误
        lines = max(1, len(text) * char_width / available_width)
        
        # 基础行高
        base_height = 16.0
        if self.method == 'gdi':
            base_height = 13.5
        elif self.method == 'pillow':
            base_height = 14.0
        
        height = lines * base_height
        
        # 更新性能统计
        self.performance_stats[self.method]['count'] += 1
        self.performance_stats[self.method]['total_time'] += 0.003
        
        return height
    
    def get_performance_stats(self) -> dict:
        """获取性能统计"""
        stats = {}
        for method, data in self.performance_stats.items():
            if data['count'] > 0:
                avg_time = data['total_time'] / data['count']
                stats[method] = {
                    'count': data['count'],
                    'total_time': data['total_time'],
                    'avg_time': avg_time,
                    'calls_per_second': 1.0 / avg_time if avg_time > 0 else 0
                }
            else:
                stats[method] = {
                    'count': 0,
                    'total_time': 0,
                    'avg_time': 0,
                    'calls_per_second': 0
                }
        return stats
    
    def set_method(self, method: str):
        """设置计算方法"""
        self.method = method

def create_mock_xlwings_range():
    """创建模拟的xlwings Range对象"""
    mock_range = MagicMock()
    mock_range.value = ""
    mock_range.row_height = 16.0
    mock_range.column_width = 10.0
    mock_range.font.size = 11
    
    def autofit_side_effect():
        # 模拟autofit行为
        text_length = len(str(mock_range.value)) if mock_range.value else 0
        mock_range.row_height = max(16.0, text_length / 20 * 16.0)
    
    mock_range.autofit.side_effect = autofit_side_effect
    return mock_range

@pytest.fixture
def test_env():
    """测试环境fixture"""
    env = TestEnvironment()
    env.setup()
    try:
        yield env
    finally:
        env.cleanup()

@pytest.fixture
def mock_archive_data():
    """模拟档案数据fixture"""
    return create_mock_archive_data()

@pytest.fixture
def mock_template_bytes():
    """模拟模板字节数据fixture"""
    return create_mock_template()

@pytest.fixture
def mock_height_calculator():
    """模拟行高计算器fixture"""
    return MockHeightCalculator()

# 性能测试装饰器
def benchmark(func):
    """性能测试装饰器"""
    def wrapper(*args, **kwargs):
        import time
        times = []
        for _ in range(TEST_CONFIG['benchmark_iterations']):
            start = time.perf_counter()
            result = func(*args, **kwargs)
            end = time.perf_counter()
            times.append(end - start)
        
        # 记录性能指标
        avg_time = sum(times) / len(times)
        min_time = min(times)
        max_time = max(times)
        
        print(f"\n性能测试结果 - {func.__name__}:")
        print(f"  平均时间: {avg_time:.4f}s")
        print(f"  最短时间: {min_time:.4f}s")
        print(f"  最长时间: {max_time:.4f}s")
        print(f"  总迭代次数: {len(times)}")
        
        return result
    return wrapper

# 跳过Windows特定测试的标记
skip_on_non_windows = pytest.mark.skipif(
    os.name != 'nt',
    reason="需要Windows环境"
)

# 跳过需要Excel的测试标记
skip_without_excel = pytest.mark.skipif(
    TEST_CONFIG['mock_excel'],
    reason="模拟模式下跳过Excel相关测试"
)