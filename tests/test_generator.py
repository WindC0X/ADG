"""
核心生成器模块单元测试
测试core/generator.py的分页算法和Excel生成逻辑
"""

import pytest
import sys
import os
import pandas as pd
import openpyxl
from unittest.mock import Mock, patch, MagicMock
from io import BytesIO

# 添加项目路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from tests.conftest import (
    benchmark,
    create_mock_archive_data,
    create_mock_template,
    MockHeightCalculator
)

class TestGeneratorUtilityFunctions:
    """测试生成器工具函数"""
    
    def test_mm_to_twip_conversion(self):
        """测试毫米到twip转换"""
        from core.generator import mm_to_twip, MM_PER_PT, TWIP
        
        # 测试已知转换
        assert mm_to_twip(0) == 0
        assert mm_to_twip(25.4) == round(25.4 / MM_PER_PT * TWIP)  # 1英寸
        
        # 测试精度
        result = mm_to_twip(10.5)
        assert isinstance(result, int)
        assert result > 0
    
    def test_pt_to_twip_conversion(self):
        """测试点到twip转换"""
        from core.generator import pt_to_twip, TWIP
        
        assert pt_to_twip(0) == 0
        assert pt_to_twip(1) == TWIP
        assert pt_to_twip(72) == 72 * TWIP  # 1英寸 = 72点
    
    def test_twip_to_pt_conversion(self):
        """测试twip到点转换"""
        from core.generator import twip_to_pt, TWIP
        
        assert twip_to_pt(0) == 0
        assert twip_to_pt(TWIP) == 1.0
        assert abs(twip_to_pt(1440) - 72.0) < 0.001  # 1英寸
    
    def test_points_to_mm_conversion(self):
        """测试点到毫米转换（向后兼容）"""
        from core.generator import points_to_mm, mm_to_points
        
        # 测试往返转换
        original_mm = 25.4
        points = mm_to_points(original_mm)
        converted_mm = points_to_mm(points)
        
        assert abs(original_mm - converted_mm) < 0.1
    
    def test_inch_to_mm_conversion(self):
        """测试英寸到毫米转换"""
        from core.generator import inch_to_mm
        
        assert inch_to_mm(1) == 25.4
        assert inch_to_mm(0) == 0
        assert abs(inch_to_mm(2.5) - 63.5) < 0.1

class TestDataLoading:
    """测试数据加载功能"""
    
    def test_load_data_xlsx(self, test_env, mock_archive_data):
        """测试加载xlsx数据"""
        from core.generator import load_data
        
        # 创建测试Excel文件
        excel_path = test_env.create_test_excel(
            mock_archive_data.to_dict('list'),
            'test_data.xlsx'
        )
        
        # 加载数据
        loaded_data = load_data(excel_path)
        
        assert loaded_data is not None
        assert isinstance(loaded_data, pd.DataFrame)
        assert len(loaded_data) == len(mock_archive_data)
        assert list(loaded_data.columns) == list(mock_archive_data.columns)
    
    @patch('core.generator.xls2xlsx')
    def test_load_data_xls_conversion(self, mock_xls2xlsx, test_env):
        """测试.xls文件自动转换"""
        from core.generator import load_data
        
        # 创建.xls文件路径
        xls_path = os.path.join(test_env.temp_dir, 'test_data.xls')
        xlsx_path = os.path.splitext(xls_path)[0] + '.xlsx'
        
        # 模拟转换函数
        mock_xls2xlsx.return_value = None
        
        # 创建转换后的xlsx文件
        mock_data = create_mock_archive_data(5)
        mock_data.to_excel(xlsx_path, index=False)
        test_env.temp_files.append(xlsx_path)
        
        # 创建空的xls文件（仅用于测试路径存在性）
        with open(xls_path, 'w') as f:
            f.write('')
        test_env.temp_files.append(xls_path)
        
        # 加载数据
        loaded_data = load_data(xls_path)
        
        # 验证转换被调用
        mock_xls2xlsx.assert_called_once_with(xls_path)
        
        # 验证数据加载成功
        assert loaded_data is not None
        assert len(loaded_data) == 5
    
    def test_load_data_nonexistent_file(self):
        """测试加载不存在文件"""
        from core.generator import load_data
        
        result = load_data('/nonexistent/path/file.xlsx')
        assert result is None
    
    def test_load_data_invalid_excel(self, test_env):
        """测试加载无效Excel文件"""
        from core.generator import load_data
        
        # 创建无效的Excel文件
        invalid_path = test_env.create_temp_file('invalid content', '.xlsx')
        
        result = load_data(invalid_path)
        assert result is None

class TestTemplateHandling:
    """测试模板处理功能"""
    
    def test_prepare_template_success(self, test_env, mock_template_bytes):
        """测试成功准备模板"""
        from core.generator import prepare_template
        
        # 创建模板文件
        template_path = os.path.join(test_env.temp_dir, 'template.xlsx')
        with open(template_path, 'wb') as f:
            f.write(mock_template_bytes)
        test_env.temp_files.append(template_path)
        
        # 准备模板
        stream = prepare_template(template_path)
        
        assert stream is not None
        assert isinstance(stream, BytesIO)
        
        # 验证流内容
        stream.seek(0)
        content = stream.read()
        assert len(content) > 0
        
        # 清理
        stream.close()
    
    def test_prepare_template_nonexistent(self):
        """测试准备不存在的模板"""
        from core.generator import prepare_template
        
        result = prepare_template('/nonexistent/template.xlsx')
        assert result is None
    
    def test_cleanup_stream(self):
        """测试流清理功能"""
        from core.generator import cleanup_stream
        
        # 创建模拟流
        stream = BytesIO(b'test content')
        stream.close = Mock()
        
        cleanup_stream(stream)
        stream.close.assert_called_once()
        
        # 测试None流
        cleanup_stream(None)  # 不应该出错

class TestGetSubset:
    """测试子集获取功能"""
    
    def test_get_subset_with_keys(self):
        """测试使用起始和结束键获取子集"""
        from core.generator import get_subset
        
        test_data = ['A', 'B', 'C', 'D', 'E', 'F']
        
        # 测试正常范围
        result = get_subset(test_data, 'B', 'E')
        assert result == ['B', 'C', 'D', 'E']
        
        # 测试只有起始键
        result = get_subset(test_data, 'C', '')
        assert result == ['C', 'D', 'E', 'F']
        
        # 测试只有结束键
        result = get_subset(test_data, '', 'D')
        assert result == ['A', 'B', 'C', 'D']
    
    def test_get_subset_invalid_keys(self):
        """测试无效键的处理"""
        from core.generator import get_subset
        
        test_data = ['A', 'B', 'C', 'D', 'E']
        
        # 测试不存在的起始键
        with patch('logging.warning') as mock_warning:
            result = get_subset(test_data, 'X', 'C')
            assert result == ['A', 'B', 'C']
            mock_warning.assert_called()
        
        # 测试不存在的结束键
        with patch('logging.warning') as mock_warning:
            result = get_subset(test_data, 'B', 'X')
            assert result == ['B', 'C', 'D', 'E']
            mock_warning.assert_called()
    
    def test_get_subset_empty_keys(self):
        """测试空键的处理"""
        from core.generator import get_subset
        
        test_data = ['A', 'B', 'C', 'D']
        
        # 测试空起始和结束键
        result = get_subset(test_data, '', '')
        assert result == test_data
        
        # 测试None键
        result = get_subset(test_data, None, None)
        assert result == test_data
    
    def test_get_subset_pandas_series(self):
        """测试处理pandas Series"""
        from core.generator import get_subset
        
        test_series = pd.Series(['A', 'B', 'C', 'D', 'E'])
        result = get_subset(test_series, 'B', 'D')
        
        assert result == ['B', 'C', 'D']

class TestColumnWidthUtilities:
    """测试列宽工具函数"""
    
    def test_get_cell_or_merged_width_simple(self):
        """测试获取简单单元格列宽"""
        from core.generator import get_cell_or_merged_width
        
        # 创建测试工作表
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 15.0
        sheet.column_dimensions['B'].width = 20.0
        
        # 测试获取列宽
        width_a = get_cell_or_merged_width(sheet, 'A1')
        width_b = get_cell_or_merged_width(sheet, 'B5')
        
        assert width_a == 15.0
        assert width_b == 20.0
    
    def test_get_cell_or_merged_width_merged(self):
        """测试获取合并单元格列宽"""
        from core.generator import get_cell_or_merged_width
        
        # 创建测试工作表
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # 设置列宽
        sheet.column_dimensions['A'].width = 10.0
        sheet.column_dimensions['B'].width = 15.0
        sheet.column_dimensions['C'].width = 20.0
        
        # 合并单元格A1:C1
        sheet.merge_cells('A1:C1')
        
        # 测试获取合并单元格总宽度
        total_width = get_cell_or_merged_width(sheet, 'A1')
        expected_width = 10.0 + 15.0 + 20.0
        
        assert total_width == expected_width
    
    def test_get_cell_or_merged_width_no_dimension(self):
        """测试未设置列宽的情况"""
        from core.generator import get_cell_or_merged_width
        
        # 创建测试工作表（不设置列宽）
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        width = get_cell_or_merged_width(sheet, 'A1')
        assert width == 0

class TestOneRowHeight:
    """测试单行高度计算"""
    
    @patch('core.generator.get_height_calculator')
    def test_get_one_row_height_xlwings(self, mock_get_calculator):
        """测试xlwings方式的单行高度计算"""
        from core.generator import get_one_row_height
        
        # 创建模拟范围
        mock_range = Mock()
        mock_range.value = None
        mock_range.row_height = 160.0  # 10行 * 16pt
        mock_range.column_width = 10.0
        mock_range.api.WrapText = False
        mock_range.autofit = Mock()
        
        # 计算单行高度
        height = get_one_row_height(mock_range)
        
        # 验证结果
        assert height == 16.0  # 160.0 / 10
        
        # 验证调用序列
        mock_range.autofit.assert_called()

class TestPaginationLogic:
    """测试分页逻辑（关键功能）"""
    
    def test_twip_precision_calculation(self):
        """测试twip精度计算的准确性"""
        from core.generator import mm_to_twip, pt_to_twip, twip_to_pt
        
        # A4纸张尺寸测试
        a4_height_mm = 297
        a4_height_twip = mm_to_twip(a4_height_mm)
        
        # 页边距测试
        margin_mm = 25.4  # 1英寸
        margin_twip = mm_to_twip(margin_mm)
        
        # 内容区域计算
        content_twip = a4_height_twip - 2 * margin_twip
        content_pt = twip_to_pt(content_twip)
        
        # 验证计算合理性
        assert a4_height_twip > 0
        assert margin_twip > 0
        assert content_twip > 0
        assert content_pt > 0
        
        # 验证精度一致性
        assert abs(twip_to_pt(pt_to_twip(72.0)) - 72.0) < 0.001

class TestArchiveDirectoryGeneration:
    """测试档案目录生成功能"""
    
    @patch('core.generator.get_height_calculator')
    @patch('openpyxl.load_workbook')
    def test_generate_one_archive_directory_basic(self, mock_load_wb, mock_get_calculator, 
                                                  test_env, mock_archive_data, mock_template_bytes):
        """测试基本的档案目录生成"""
        from core.generator import generate_one_archive_directory
        
        # 设置模拟
        mock_calculator = MockHeightCalculator()
        mock_get_calculator.return_value = mock_calculator
        
        # 创建模拟工作簿
        mock_wb = Mock()
        mock_sheet = Mock()
        mock_wb.worksheets = [mock_sheet]
        mock_wb.save = Mock()
        mock_wb.close = Mock()
        mock_load_wb.return_value = mock_wb
        
        # 设置工作表属性
        mock_sheet.page_setup.paperSize = "9"  # A4
        mock_sheet.page_setup.orientation = 'portrait'
        mock_sheet.page_setup.scale = 100
        mock_sheet.page_margins.top = 1.0
        mock_sheet.page_margins.bottom = 1.0
        mock_sheet.page_margins.footer = 0.5
        mock_sheet.max_column = 4
        mock_sheet.row_dimensions = {}
        mock_sheet.column_dimensions = {}
        mock_sheet.merged_cells.ranges = []
        mock_sheet.row_breaks = Mock()
        mock_sheet.row_breaks.append = Mock()
        
        # 设置行维度
        for i in range(1, 20):
            mock_row_dim = Mock()
            mock_row_dim.height = 16.0
            mock_sheet.row_dimensions[i] = mock_row_dim
        
        # 设置单元格
        mock_cells = {}
        for row in range(1, 20):
            for col in range(1, 5):
                cell = Mock()
                cell.font = Mock()
                cell.border = Mock()
                cell.fill = Mock()
                cell.alignment = Mock()
                cell.number_format = 'General'
                cell.protection = Mock()
                cell.value = None
                mock_cells[(row, col)] = cell
        
        def get_cell(row, column):
            return mock_cells.get((row, column), Mock())
        
        mock_sheet.cell = get_cell
        
        # 创建模板流
        template_stream = BytesIO(mock_template_bytes)
        
        # 创建模拟xlwings范围
        mock_range = Mock()
        mock_range.value = ""
        mock_range.row_height = 16.0
        mock_range.column_width = 10.0
        mock_range.font.size = 11
        mock_range.autofit = Mock()
        
        # 执行生成
        pages = generate_one_archive_directory(
            archive_data=mock_archive_data.head(5),  # 只用前5条数据
            template_stream=template_stream,
            output_folder=test_env.temp_dir,
            archive_id="TEST001",
            rng_for_calc=mock_range,
            index=1,
            column_mapping={1: '案卷档号', 2: '文件名', 3: '页数', 4: '备注'},
            autofit_columns=[2, 4],
            static_cells={'A1': '测试目录'},
            title_row_num=4,
            page_height_mm=297
        )
        
        # 验证结果
        assert pages >= 1
        mock_wb.save.assert_called_once()
        mock_wb.close.assert_called_once()
    
    @patch('core.generator.get_height_calculator')
    def test_generate_with_print_service(self, mock_get_calculator, test_env, mock_archive_data):
        """测试带打印服务的生成"""
        from core.generator import generate_one_archive_directory
        
        mock_calculator = MockHeightCalculator()
        mock_get_calculator.return_value = mock_calculator
        
        # 模拟模板流
        template_stream = BytesIO(create_mock_template())
        
        # 模拟xlwings范围
        mock_range = Mock()
        mock_range.autofit = Mock()
        mock_range.font.size = 11
        
        with patch('openpyxl.load_workbook') as mock_load_wb, \
             patch('core.generator.get_print_service') as mock_get_print_service:
            
            # 设置模拟工作簿
            mock_wb = Mock()
            mock_sheet = Mock()
            mock_wb.worksheets = [mock_sheet]
            mock_wb.save = Mock()
            mock_wb.close = Mock()
            mock_load_wb.return_value = mock_wb
            
            # 基本工作表设置
            mock_sheet.page_setup.paperSize = "9"
            mock_sheet.page_setup.orientation = 'portrait'
            mock_sheet.page_setup.scale = 100
            mock_sheet.page_margins.top = 1.0
            mock_sheet.page_margins.bottom = 1.0
            mock_sheet.page_margins.footer = 0.5
            mock_sheet.max_column = 4
            mock_sheet.row_dimensions = {}
            mock_sheet.merged_cells.ranges = []
            mock_sheet.row_breaks = Mock()
            
            # 设置打印服务
            mock_print_service = Mock()
            mock_print_service.async_print = Mock()
            mock_get_print_service.return_value = mock_print_service
            
            # 执行生成（启用打印）
            pages = generate_one_archive_directory(
                archive_data=mock_archive_data.head(3),
                template_stream=template_stream,
                output_folder=test_env.temp_dir,
                archive_id="PRINT_TEST",
                rng_for_calc=mock_range,
                index=1,
                column_mapping={1: '案卷档号', 2: '文件名'},
                autofit_columns=[2],
                direct_print=True,
                printer_name="Test Printer",
                print_copies=2
            )
            
            # 验证打印服务被调用
            mock_print_service.async_print.assert_called_once()

if __name__ == "__main__":
    pytest.main([__file__])